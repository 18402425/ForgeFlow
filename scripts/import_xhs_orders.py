#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - friendly CLI failure
    load_workbook = None


ORDER_FIELDS = [
    "snapshot_date",
    "order_id",
    "source_order_ref",
    "manual_sequence",
    "order_created_at",
    "due_at",
    "manual_actual_ship_at",
    "sku_id",
    "sku_name",
    "sku_spec",
    "quantity",
    "revenue_estimate",
    "profit_estimate",
    "platform_tags",
    "manual_shortage_flag",
    "manual_equipment_changed",
    "notes",
]

FIELD_CANDIDATES = {
    "order_ref": ["订单号", "订单编号", "订单ID"],
    "order_status": ["订单状态", "售后状态"],
    "sku_name": ["SKU名称", "商品名称", "商品标题"],
    "sku_spec": ["SKU规格", "商品规格", "规格", "规格名称"],
    "quantity": ["SKU件数", "购买数量", "数量", "商品数量"],
    "created_at": ["支付时间", "订单创建时间", "下单时间", "创建时间"],
    "due_at": ["承诺发货时间", "预售订单截止发货时间", "最晚发货时间"],
    "ship_at": ["订单发货时间", "发货时间"],
    "finished_at": ["订单完成时间", "完成时间"],
    "revenue": ["商家应收金额(元)（支付金额）", "商家应收金额", "支付金额", "订单金额"],
    "tags": ["订单标记", "订单类型", "异常原因"],
}


def normalize_text(value: object) -> str:
    text = str(value or "").strip()
    text = text.replace("\n", " ")
    return re.sub(r"\s+", " ", text)


def normalize_key(*parts: object) -> str:
    return "||".join(normalize_text(part).lower() for part in parts)


def mask_order_ref(order_ref: str) -> str:
    text = normalize_text(order_ref)
    if len(text) <= 8:
        return text
    return f"{text[:4]}...{text[-4:]}"


def get_first(row: dict[str, object], logical_name: str) -> str:
    for field in FIELD_CANDIDATES[logical_name]:
        if field in row and normalize_text(row.get(field)):
            return normalize_text(row.get(field))
    return ""


def split_tags(row: dict[str, object]) -> list[str]:
    tags: list[str] = []
    for field in FIELD_CANDIDATES["tags"]:
        text = normalize_text(row.get(field))
        if not text:
            continue
        tags.extend([part.strip() for part in re.split(r"[,，|/]", text) if part.strip()])
    return tags


def load_xlsx_rows(source_xlsx: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required. Install it with: python3 -m pip install openpyxl")
    workbook = load_workbook(source_xlsx, data_only=True, read_only=False)
    sheet = workbook.active
    headers = [normalize_text(cell.value) for cell in sheet[1]]
    rows: list[dict[str, object]] = []
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        if not any(normalize_text(value) for value in raw):
            continue
        rows.append({headers[index]: raw[index] for index in range(len(headers))})
    return rows, {"sheet_name": sheet.title, "row_count": len(rows), "column_count": len(headers)}


def load_csv(path: Path) -> list[dict[str, str]]:
    if not path or not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [{key: (value or "").strip() for key, value in row.items()} for row in csv.DictReader(handle)]


def build_sku_lookup(sku_rows: list[dict[str, str]], alias_rows: list[dict[str, str]]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for row in sku_rows:
        key = normalize_key(row.get("sku_name", ""), row.get("sku_spec", ""))
        if key.strip("|"):
            lookup[key] = row.get("sku_id", "")
    for row in alias_rows:
        key = normalize_key(row.get("platform_product_name", ""), row.get("platform_sku_spec", ""))
        sku_id = row.get("sku_id", "")
        if key.strip("|") and sku_id:
            lookup[key] = sku_id
    return lookup


def write_csv(path: Path, rows: list[dict[str, str]], headers: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def should_include_status(status: str, status_filter: str) -> bool:
    if not status_filter:
        return status not in {"已取消", "已关闭", "退款成功", "交易关闭"}
    allowed = {item.strip() for item in status_filter.split("|") if item.strip()}
    return status in allowed


def convert_rows(
    rows: list[dict[str, object]],
    sku_lookup: dict[str, str],
    snapshot_date: str,
    order_id_prefix: str,
    status_filter: str,
) -> tuple[list[dict[str, str]], list[dict[str, str]], dict[str, object]]:
    converted: list[dict[str, str]] = []
    unknown: list[dict[str, str]] = []
    status_counts: Counter[str] = Counter()
    unknown_pairs: Counter[str] = Counter()

    for index, row in enumerate(rows, start=1):
        status = get_first(row, "order_status")
        status_counts[status or "UNKNOWN"] += 1
        if not should_include_status(status, status_filter):
            continue

        sku_name = get_first(row, "sku_name")
        sku_spec = get_first(row, "sku_spec")
        sku_id = sku_lookup.get(normalize_key(sku_name, sku_spec), "")
        order_id = f"{order_id_prefix}-{index:03d}"

        if not sku_id:
            unknown_pairs[f"{sku_name} || {sku_spec}"] += 1
            unknown.append(
                {
                    "order_id": order_id,
                    "platform_product_name": sku_name,
                    "platform_sku_spec": sku_spec,
                    "count": "1",
                }
            )
            continue

        tags = split_tags(row)
        converted.append(
            {
                "snapshot_date": snapshot_date,
                "order_id": order_id,
                "source_order_ref": mask_order_ref(get_first(row, "order_ref")),
                "manual_sequence": str(index),
                "order_created_at": get_first(row, "created_at"),
                "due_at": get_first(row, "due_at"),
                "manual_actual_ship_at": get_first(row, "ship_at") or get_first(row, "finished_at"),
                "sku_id": sku_id,
                "sku_name": sku_name,
                "sku_spec": sku_spec,
                "quantity": get_first(row, "quantity") or "1",
                "revenue_estimate": get_first(row, "revenue"),
                "profit_estimate": "",
                "platform_tags": "|".join(tags),
                "manual_shortage_flag": "true" if any("缺货" in tag for tag in tags) else "false",
                "manual_equipment_changed": "false",
                "notes": f"订单状态={status or 'UNKNOWN'}",
            }
        )

    profile = {
        "raw_order_count": len(rows),
        "converted_order_count": len(converted),
        "unknown_sku_count": len(unknown),
        "status_counts": dict(status_counts),
        "top_unknown_pairs": [
            {"sku_pair": pair, "count": count} for pair, count in unknown_pairs.most_common(20)
        ],
    }
    return converted, unknown, profile


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import Xiaohongshu order Excel into ForgeFlow standard orders.csv.")
    parser.add_argument("--source-xlsx", type=Path, required=True)
    parser.add_argument("--sku", type=Path, default=Path("examples/standard/sku_catalog.csv"))
    parser.add_argument("--aliases", type=Path, default=Path("examples/standard/sku_aliases.csv"))
    parser.add_argument("--output", type=Path, default=Path("outputs/imported/orders.csv"))
    parser.add_argument("--unknown-output", type=Path, default=Path("outputs/imported/unknown_sku.csv"))
    parser.add_argument("--profile-output", type=Path, default=Path("outputs/imported/import_profile.json"))
    parser.add_argument("--snapshot-date", required=True)
    parser.add_argument("--order-id-prefix", default="")
    parser.add_argument("--status-filter", default="待配货", help="Pipe-separated status allowlist. Empty means all non-cancelled rows.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        raw_rows, workbook_meta = load_xlsx_rows(args.source_xlsx)
    except Exception as error:
        print(f"Import failed: {error}", file=sys.stderr)
        return 2

    order_id_prefix = args.order_id_prefix or f"XHS{args.snapshot_date.replace('-', '')}"
    converted, unknown, profile = convert_rows(
        raw_rows,
        build_sku_lookup(load_csv(args.sku), load_csv(args.aliases)),
        args.snapshot_date,
        order_id_prefix,
        args.status_filter,
    )
    profile["source_xlsx"] = str(args.source_xlsx)
    profile["workbook"] = workbook_meta
    profile["sku_catalog"] = str(args.sku)
    profile["aliases"] = str(args.aliases)

    write_csv(args.output, converted, ORDER_FIELDS)
    write_csv(args.unknown_output, unknown, ["order_id", "platform_product_name", "platform_sku_spec", "count"])
    args.profile_output.parent.mkdir(parents=True, exist_ok=True)
    args.profile_output.write_text(json.dumps(profile, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(
        json.dumps(
            {
                "orders": str(args.output),
                "unknown_sku": str(args.unknown_output),
                "profile": str(args.profile_output),
                **profile,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 1 if unknown else 0


if __name__ == "__main__":
    raise SystemExit(main())
