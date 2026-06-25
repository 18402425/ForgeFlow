import csv
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = Path("/Users/kellenwang/Downloads/小红书订单查询2026-06-08-15_11_383f4ff8（全部都是未完成订单）.xlsx")
ORDERS_20_CSV = ROOT / "fixtures" / "orders_20.csv"
SKU_CSV = ROOT / "fixtures" / "sku.csv"
OUTPUT_XLSX = ROOT / "reports" / "p0a_used_20_orders.xlsx"


SOURCE_FIELDS = [
    "订单号",
    "订单状态",
    "订单类型",
    "订单标记",
    "SKU名称",
    "SKU规格",
    "SKU件数",
    "商品总价(元)",
    "商家应收金额(元)（支付金额）",
    "订单创建时间",
    "承诺发货时间",
    "订单发货时间",
    "订单完成时间",
    "商品名称",
    "商品ID",
    "规格ID",
]


def read_orders_20():
    with ORDERS_20_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_skus():
    with SKU_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        return {row["sku_id"]: row for row in csv.DictReader(handle)}


def parse_dt(value):
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")


def fmt_dt(value):
    return "" if value is None else value.strftime("%Y-%m-%d %H:%M:%S")


def as_float(value):
    return float(value or 0)


def process_hours(p0a_row, sku):
    return (
        as_float(sku["standard_print_hours"]) * as_float(p0a_row["quantity"])
        + (as_float(sku["post_process_minutes"]) + as_float(sku["package_minutes"])) / 60
    )


def enrich_manual_baseline(p0a_rows, skus):
    enriched = []
    for row in p0a_rows:
        sku = skus[row["sku_id"]]
        ship_at = parse_dt(row["manual_actual_ship_at"])
        inferred_start = ship_at - timedelta(hours=process_hours(row, sku)) if ship_at else None
        enriched.append({**row, "manual_inferred_start_at": inferred_start})
    ordered = sorted(
        enriched,
        key=lambda row: (
            row["manual_inferred_start_at"] or datetime.max,
            parse_dt(row["manual_actual_ship_at"]) or datetime.max,
            int(row["manual_sequence"] or 9999),
        ),
    )
    rank_by_order_id = {row["order_id"]: index for index, row in enumerate(ordered, start=1)}
    for row in enriched:
        row["manual_baseline_rank"] = rank_by_order_id[row["order_id"]]
    return enriched


def read_source_rows():
    workbook = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    sheet = workbook["包裹详情"]
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for excel_row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(headers, values))
        row["原始Excel行号"] = excel_row_number
        rows.append(row)
    return rows


def norm(value):
    return "" if value is None else str(value).strip()


def find_source_row(p0a_row, source_rows, used_excel_rows):
    source_ref = norm(p0a_row["source_order_ref"])
    sku_spec = norm(p0a_row["sku_spec"])
    due_at = norm(p0a_row["due_at"])
    ship_at = norm(p0a_row["manual_actual_ship_at"])

    candidates = []
    for row in source_rows:
        if row["原始Excel行号"] in used_excel_rows:
            continue
        if not norm(row["订单号"]).endswith(source_ref):
            continue
        if norm(row["SKU规格"]) != sku_spec:
            continue
        if due_at and norm(row["承诺发货时间"]) != due_at:
            continue
        if ship_at and norm(row["订单发货时间"]) != ship_at:
            continue
        candidates.append(row)

    if candidates:
        return candidates[0]

    # Fallback for source rows whose SKU text is slightly normalized in orders_20.csv.
    for row in source_rows:
        if row["原始Excel行号"] in used_excel_rows:
            continue
        if norm(row["订单号"]).endswith(source_ref) and norm(row["承诺发货时间"]) == due_at:
            return row
    return None


def style_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column in enumerate(sheet.columns, start=1):
        max_len = 0
        for cell in column:
            max_len = max(max_len, len(norm(cell.value)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 36)


def main():
    p0a_rows = enrich_manual_baseline(read_orders_20(), read_skus())
    source_rows = read_source_rows()
    used_excel_rows = set()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "P0a使用的20条"

    headers = [
        "P0a序号",
        "P0a订单ID",
        "原始Excel行号",
        "原始订单号",
        "订单号后6位",
        "选择说明",
        "P0a SKU ID",
        "P0a SKU名称",
        "P0a SKU规格",
        "旧人工序号",
        "新人工基线排序",
        "倒推生产开始时间",
        "P0a决策日期",
        "P0a预计收入",
        "P0a预计毛利",
        *SOURCE_FIELDS[1:],
    ]
    sheet.append(headers)

    for index, p0a_row in enumerate(p0a_rows, start=1):
        source_row = find_source_row(p0a_row, source_rows, used_excel_rows)
        if source_row:
            used_excel_rows.add(source_row["原始Excel行号"])
        else:
            source_row = {field: "" for field in SOURCE_FIELDS}
            source_row["原始Excel行号"] = ""

        note = "有承诺发货时间和历史发货时间，可倒推 P0a 人工基线生产开始时间；不包含客户姓名、电话、地址。"
        if p0a_row["notes"] == "multi_sku_source_order":
            note = "同一平台订单拆成多 SKU 行；按 SKU 行进入 P0a 回测。"

        sheet.append(
            [
                index,
                p0a_row["order_id"],
                source_row["原始Excel行号"],
                norm(source_row["订单号"]),
                p0a_row["source_order_ref"],
                note,
                p0a_row["sku_id"],
                p0a_row["sku_name"],
                p0a_row["sku_spec"],
                p0a_row["manual_sequence"],
                p0a_row["manual_baseline_rank"],
                fmt_dt(p0a_row["manual_inferred_start_at"]),
                p0a_row["snapshot_date"],
                p0a_row["revenue_estimate"],
                p0a_row["profit_estimate"],
                *[norm(source_row.get(field)) for field in SOURCE_FIELDS[1:]],
            ]
        )

    note_sheet = workbook.create_sheet("说明")
    note_sheet.append(["项目", "说明"])
    note_sheet.append(["这 20 条是不是原始 Excel 前 20 行", "不是。它们是从原始订单中挑出的 P0a 回测样本。"])
    note_sheet.append(["选择标准", "优先选择同时具备承诺发货时间、历史发货时间、SKU 可映射、且不需要保留 PII 的订单行。"])
    note_sheet.append(["为什么需要历史发货时间", "P0a 当前没有真实人工计划日志，所以用历史发货时间倒推候选生产开始时间。"])
    note_sheet.append(["人工基线排序方法", "历史发货时间 - SKU标准打印时长 - 后处理时间 - 包装时间，得到倒推生产开始时间；再按该时间从早到晚排序。"])
    note_sheet.append(["资源约束回放", "排序后仍需用耗材库存、推荐设备、设备可用时段回放；缺料、设备容量不足、晚于承诺发货时间会进入风险或阻断。"])
    note_sheet.append(["隐私处理", "本文件不包含收件人姓名、电话、地址、身份证等 PII 字段。"])
    note_sheet.append(["完整订单号", "为方便你核对原始数据，本文件保留了原始订单号；对外展示时建议只用订单号后 6 位。"])
    note_sheet.append(["后续建议", "下一步应增加全量 214 条导入体检，输出可回测、不可回测及原因归类。"])

    style_sheet(sheet)
    style_sheet(note_sheet)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
