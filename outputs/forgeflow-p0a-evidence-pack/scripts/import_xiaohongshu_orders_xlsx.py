#!/usr/bin/env python3

from __future__ import annotations

import csv
import argparse
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FIXTURES_DIR = ROOT / "fixtures"
REPORTS_DIR = ROOT / "reports"
SOURCE_XLSX = Path("/Users/kellenwang/Downloads/小红书订单查询2026-06-08-15_11_383f4ff8.xlsx")
SNAPSHOT_DATE = "2026-06-08"

FULL_OUTPUT = FIXTURES_DIR / "orders_xhs_20260608_full_mapped.csv"
PENDING_OUTPUT = FIXTURES_DIR / "orders_xhs_20260608_pending_7.csv"
PROFILE_OUTPUT = REPORTS_DIR / "xhs_20260608_import_profile.json"

ORDER_FIELDS = [
    "snapshot_date",
    "order_id",
    "source_order_ref",
    "manual_sequence",
    "order_created_at",
    "due_at",
    "manual_actual_ship_at",
    "sku_id",
    "sku_name",
    "sku_spec",
    "quantity",
    "revenue_estimate",
    "profit_estimate",
    "platform_tags",
    "manual_shortage_flag",
    "manual_equipment_changed",
    "notes",
]


def normalize_text(value: object) -> str:
    text = str(value or "").strip()
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def mask_order_id(order_id: str) -> str:
    text = normalize_text(order_id)
    if len(text) <= 8:
        return text
    return f"{text[:4]}...{text[-4:]}"


def parse_money(value: object) -> str:
    text = normalize_text(value)
    return text if text else ""


def split_tags(*parts: object) -> list[str]:
    tags: list[str] = []
    for part in parts:
        text = normalize_text(part)
        if not text:
            continue
        tags.extend([item.strip() for item in text.split(",") if item.strip()])
    return tags


def infer_sku_id(sku_name: str, sku_spec: str) -> str:
    combo = f"{normalize_text(sku_name)} {normalize_text(sku_spec)}"

    if "绒仙子" in combo or "白模" in combo:
        if "12cm*8cm*15cm高" in combo:
            return "RX_WHITE_15"
        if "8cm*5cm*10cm高" in combo:
            return "RX_WHITE_10"

    if "鸭吉吉" in combo or "鸭喆" in combo:
        if "白模，8cm x 8cm x 10cm" in combo:
            return "JJ_WHITE_8"
        if "黑色炫彩" in combo and "12cm x 12cm x 15cm" in combo:
            return "JJ_BLACK_12"
        if "黑色炫彩" in combo and "8cm x 8cm x 10cm" in combo:
            return "JJ_BLACK_8"
        if "粉色炫彩" in combo and "12cm x 12cm x 15cm" in combo:
            return "JJ_PINK_12"
        if "粉色炫彩" in combo and "8cm x 8cm x 10cm" in combo:
            return "JJ_PINK_8"
        if "黄色原皮" in combo and "12cm x 12cm x 15cm" in combo:
            return "JJ_YELLOW_12"
        if "黄色原皮" in combo and "6cm x 6cm x 8cm" in combo:
            return "JJ_YELLOW_6"
        if ("绿色炫彩" in combo or "青色炫彩" in combo) and "12cm x 12cm x 15cm" in combo:
            return "JJ_GREEN_12"
        if ("绿色炫彩" in combo or "青色炫彩" in combo) and "8cm x 8cm x 10cm" in combo:
            return "JJ_GREEN_8"
        if "鸭蓝蓝" in combo and "8cm x 8cm x 10cm" in combo:
            return "JJ_BLUE_8"
        if "8cm x 8cm x 10cm" in combo:
            return "JJ_COLOR_PENDING_8"

    if "大耳帽兜" in combo:
        if "15cm高（包含蛋）" in combo:
            return "DH_HAT_15_EGG"
        if "15cm高（不包含蛋）" in combo:
            return "DH_HAT_15_NO_EGG"
        if "10cm（包含蛋）" in combo:
            return "DH_HAT_10_EGG"
        if "10cm（不包含蛋）" in combo:
            return "DH_HAT_10_NO_EGG"
        if "6cm，只有蛋" in combo:
            return "DH_HAT_EGG_6"

    if "恶魔叮彩色编织绳钥匙扣" in combo:
        return "EMD_KEY_BLACK"

    if "恶魔叮" in combo and "钥匙扣" in combo:
        return "EMD_FIGURE_KEY_BLACK"

    if "翠腚夫人" in combo:
        if "10cm*10cm*15cm" in combo:
            return "CDFR_15"
        if "7cm*7cm*10cm" in combo:
            return "CDFR_10"

    if "恶魔狼" in combo:
        if "紫色炫彩" in combo or "7cm*20cm*16cm" in combo:
            return "EMWOLF_PURPLE_16"
        return "EMWOLF_10"

    if "雷霆粑粑干" in combo:
        return "LTBBG_8"

    return ""


def load_rows(source_xlsx: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(source_xlsx, data_only=True, read_only=False)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        rows.append({headers[index]: row[index] for index in range(len(headers))})
    return rows, {"sheet_name": sheet.title, "row_count": len(rows), "column_count": len(headers)}


def convert_rows(
    rows: list[dict[str, object]],
    source_xlsx: Path,
    snapshot_date: str,
    order_id_prefix: str,
) -> tuple[list[dict[str, str]], list[dict[str, str]], dict[str, object]]:
    mapped_rows: list[dict[str, str]] = []
    pending_rows: list[dict[str, str]] = []
    status_counter: Counter[str] = Counter()
    mapped_status_counter: Counter[str] = Counter()
    unmapped_pairs: Counter[str] = Counter()

    for index, row in enumerate(rows, start=1):
        order_status = normalize_text(row.get("订单状态"))
        sku_name = normalize_text(row.get("SKU名称"))
        sku_spec = normalize_text(row.get("SKU规格"))
        sku_id = infer_sku_id(sku_name, sku_spec)
        status_counter[order_status] += 1

        if not sku_id:
            unmapped_pairs[f"{sku_name} || {sku_spec}"] += 1
            continue

        mapped_status_counter[order_status] += 1
        tags = split_tags(row.get("订单标记"), row.get("订单类型"))
        shortage_flag = "true" if ("缺货" in tags or "缺货" in normalize_text(row.get("异常原因"))) else "false"
        normalized = {
            "snapshot_date": snapshot_date,
            "order_id": f"{order_id_prefix}-{index:03d}",
            "source_order_ref": mask_order_id(normalize_text(row.get("订单号"))),
            "manual_sequence": str(index),
            "order_created_at": normalize_text(row.get("支付时间") or row.get("订单创建时间")),
            "due_at": normalize_text(row.get("承诺发货时间") or row.get("预售订单截止发货时间")),
            "manual_actual_ship_at": normalize_text(row.get("订单发货时间") or row.get("订单完成时间")),
            "sku_id": sku_id,
            "sku_name": sku_name,
            "sku_spec": sku_spec,
            "quantity": normalize_text(row.get("SKU件数") or "1"),
            "revenue_estimate": parse_money(row.get("商家应收金额(元)（支付金额）")),
            "profit_estimate": "",
            "platform_tags": "|".join(tags),
            "manual_shortage_flag": shortage_flag,
            "manual_equipment_changed": "false",
            "notes": f"订单状态={order_status or '未知'}",
        }
        mapped_rows.append(normalized)
        if order_status == "待配货":
            pending_rows.append(normalized)

    profile = {
        "source_xlsx": str(source_xlsx),
        "snapshot_date": snapshot_date,
        "raw_order_count": len(rows),
        "mapped_order_count": len(mapped_rows),
        "pending_order_count": len(pending_rows),
        "mapping_coverage_pct": round(len(mapped_rows) * 100 / max(1, len(rows)), 2),
        "status_counts": dict(status_counter),
        "mapped_status_counts": dict(mapped_status_counter),
        "top_unmapped_pairs": [
            {"sku_pair": pair, "count": count} for pair, count in unmapped_pairs.most_common(12)
        ],
    }
    return mapped_rows, pending_rows, profile


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=ORDER_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="导入小红书订单 Excel，输出 ForgeFlow P0a CSV。")
    parser.add_argument("--source-xlsx", type=Path, default=SOURCE_XLSX)
    parser.add_argument("--snapshot-date", default=SNAPSHOT_DATE)
    parser.add_argument("--full-output", type=Path, default=FULL_OUTPUT)
    parser.add_argument("--pending-output", type=Path, default=PENDING_OUTPUT)
    parser.add_argument("--profile-output", type=Path, default=PROFILE_OUTPUT)
    parser.add_argument("--order-id-prefix", default="")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    order_id_prefix = args.order_id_prefix or f"XHS{args.snapshot_date.replace('-', '')[4:]}"
    rows, workbook_meta = load_rows(args.source_xlsx)
    mapped_rows, pending_rows, profile = convert_rows(
        rows,
        args.source_xlsx,
        args.snapshot_date,
        order_id_prefix,
    )
    profile["workbook_meta"] = workbook_meta
    write_csv(args.full_output, mapped_rows)
    write_csv(args.pending_output, pending_rows)
    args.profile_output.parent.mkdir(parents=True, exist_ok=True)
    args.profile_output.write_text(json.dumps(profile, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "full_output": str(args.full_output),
        "pending_output": str(args.pending_output),
        "profile_output": str(args.profile_output),
        "mapped_order_count": len(mapped_rows),
        "pending_order_count": len(pending_rows),
        "coverage_pct": profile["mapping_coverage_pct"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
